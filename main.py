from tracemalloc import start
from urllib import response
from openpyxl import load_workbook
import requests
import bs4
from bs4 import BeautifulSoup
import pandas as pd
import xlsxwriter

def First5():
    outWorkbook = xlsxwriter.Workbook('f2s3.xlsx')
    outSheet = outWorkbook.add_worksheet()
    outSheet.write('A1', 'Players')
    outSheet.write('B1', 'Team')
    outSheet.write('C1', 'Home Wins')
    outSheet.write('D1', 'Away Wins')
    outSheet.write('E1', 'Method Score')

    #team = ['cha','det','hou','ind',"utah",'orl','gs','mem','bkn','mil','ny','sac','chi','den','okc','phx','tor','lac']
    team = ['den']
    #team = ['wsh','tor','atl','chi','mem','okc','phx','det','por','cle','ind','mia','phi','gs','lac','lal','den','min','sac','bkn','orl','no','hou','dal','bos','utah','cha','sa','ny','mil']
    item = 0
    num_Play = 2
    lst = []
    while item < len(team):
        print(team[item])
        # Get the game
        nba_Games = 'https://www.espn.com/nba/team/schedule/_/name/' + team[item]
        response = requests.get(nba_Games)
        soup = BeautifulSoup(response.text, 'lxml')
        start_Game = soup.find_all('a', {'class':'AnchorLink'})[num_Play].get('href')
        while True:
            try:
                # Add game to the list
                #print(start_Game)
                if 'https://www.espn.com/nba/game/' in start_Game:
                    lst.append(start_Game)            
                    text_score = soup.find_all('a', {'class':'AnchorLink'})[num_Play].text
                    if '1:00 PM' in text_score or '5:00 PM' in text_score or '3:00 PM' in text_score or '3:30 PM' in text_score or '6:00 PM' in text_score or '6:30 PM' in text_score or '7:00 PM' in text_score or '7:30 PM' in text_score or '8:00 PM' in text_score or '9:00 PM' in text_score or '9:30 PM' in text_score or '10:00 PM' in text_score or '8:30 PM' in text_score or '10:30 PM' in text_score:
                        break
                num_Play = 1 + num_Play
                start_Game = soup.find_all('a', {'class':'AnchorLink'})[num_Play].get('href')
            except:
                break
        # Get the first scorer
        lst.pop()
        i = 0
        # Check each score and determine the first player 
        while i < len(lst):
            try:
                game_Play = 0
                game_URL = lst[i].replace('/game/','/playbyplay/')
                response = requests.get(game_URL)
                soup = BeautifulSoup(response.text, 'lxml')
                if 'https://www.espn.com/nba/playbyplay/' in game_URL:
                    #the_Score = str(soup.find_all('td', {'class':"combined-score"})[game_Play].text)
                    the_Player = str(soup.find_all('td', {'class':"playByPlay__text tl Table__TD"})[game_Play].text)
                    teamA = str(soup.find_all('tr', {'class':"Table__TR Table__TR--sm Table__even"})[0].text)[:4]
                    teamH = str(soup.find_all('tr', {'class':"Table__TR Table__TR--sm Table__even"})[1].text)[:4]
                    score_Home = str(soup.find_all('tr', {'class':"playByPlay__tableRow fw-bold Table__TR Table__TR--sm Table__even"})[0].text)[-1]
                    score_Away = str(soup.find_all('tr', {'class':"playByPlay__tableRow fw-bold Table__TR Table__TR--sm Table__even"})[0].text)[-2]
                    # Remove the number off the team name
                    awayTeam = ''.join((x for x in teamA if not x.isdigit()))
                    homeTeam = ''.join((x for x in teamH if not x.isdigit()))
                # Check if either of the scores is greater than 0
                while True:
                    if int(score_Home) > 0:
                        winTeam = homeTeam
                        break
                    elif int(score_Away) > 0:
                        winTeam = awayTeam
                        break
                # Winning Player
                the_Player = str(soup.find_all('td', {'class':"playByPlay__text tl clr-btn Table__TD"})[0].text)
                player = the_Player.split()[0] + " " + the_Player.split()[1]
                methodScore = the_Player
                #print('The First Basket is ' + player)
                #print(team[item])
                nameTeam = team[item]
                #print(nameTeam)
                #print(winTeam)
                if winTeam in homeTeam and winTeam.lower() == nameTeam:
                    winner = 'h'                   
                    # Place 1 point under Home Category
                    # Place 1 point for this player
                elif winTeam in awayTeam and winTeam.lower() == nameTeam:
                    winner = 'a'      
                    # Place 1 point under Away Category
                    # Place 1 point for this player
                # If the player team matches the original team than add it to the excel sheet
                if winTeam.lower() == nameTeam:
                    outSheet.write(i + 1, 0, str(player))
                    outSheet.write(i + 1, 1, str(winTeam))
                    if winner == 'h':
                        outSheet.write(i + 1, 2, winner)
                        outSheet.write(i + 1, 4, methodScore)
                    else:
                        outSheet.write(i + 1, 3, winner)
                        outSheet.write(i + 1, 4, methodScore)
                i = i + 1
            except:
                break
        item = item + 1
        num_Play = 2
        game_Play = 0
        i = 0
    outWorkbook.close()
    return


def PlayerPropPTS():
    outWorkbook = xlsxwriter.Workbook('f2s6.xlsx')
    outSheet = outWorkbook.add_worksheet()
    outSheet.write('A1', 'Players')
    outSheet.write('B1', 'PPG')
    outSheet.write('C1', '2PT FG')
    outSheet.write('D1', '2PT FGA')
    outSheet.write('E1', '3PT FG')
    outSheet.write('F1', '3PT FGA')
    outSheet.write('G1', 'FT')
    outSheet.write('H1', 'FTA')
    outSheet.write('I1', 'MONTH')
    outSheet.write('J1', 'PLACE')

    playerName = "Darius Garland"
    oppPts = float(22.8)
    opp2PTFGA = float(0.451)
    oppFTA = float(0.831)
    opp3pm = float(2.5)
    num_Play = 0

    site = "https://www.basketball-reference.com/leagues/NBA_2023_per_game.html"
    response = requests.get(site)
    soup = BeautifulSoup(response.text,'lxml')
    
    while True:
        try:
            check_Player = soup.find_all('td', {'data-stat':'player'})[num_Play].text
            
            #print(check_Player)
            #print(num_Play)
            getLink = soup.find_all('td', {'data-stat':'player'})[num_Play].find('a').get('href')
            theLink = getLink.replace("/players/","https://www.basketball-reference.com/players/")
            #print(theLink)
            if playerName == check_Player:

                # Player Shooting
                pt2 = float(soup.find_all('td', {'data-stat':'fg2_per_g'})[num_Play].text)
                pt2a = float(soup.find_all('td', {'data-stat':'fg2a_per_g'})[num_Play].text)
                pt3 = float(soup.find_all('td', {'data-stat':'fg3_per_g'})[num_Play].text)
                #pt3a = soup.find_all('td', {'data-stat':'fg3a_per_g'})[num_Play].text
                ft = float(soup.find_all('td', {'data-stat':'ft_per_g'})[num_Play].text)
                fta = float(soup.find_all('td', {'data-stat':'fta_per_g'})[num_Play].text)

                theLink = getLink.replace("/players/","https://www.basketball-reference.com/players/")
                splitCheck = theLink.replace(".html","/splits/2023")
                response = requests.get(splitCheck)
                soup = BeautifulSoup(response.text,'lxml')

                # Current Averages
                ppg = float(soup.find_all('td',{'data-stat':'pts_per_g'})[0].text)
                #print(ppg)

                # Home Average
                ppgH = float(soup.find_all('td',{'data-stat':'pts_per_g'})[1].text)
                #print(ppgH)

                # Away Average
                ppgA = float(soup.find_all('td',{'data-stat':'pts_per_g'})[2].text)
                #print(ppgA)

                # Defense Stats
                print('')
                print(check_Player)
                #oppPts = float(input("Opponent Points: "))
                #opp3pm = float(input("Opponent 3 Points: "))
                
                #loca = input("Opponent Game Location: ")

                # Projections
                expPTS = float(0)
                expPTSH = float(0)
                expPTSA = float(0)

                if ppg - oppPts <= -4:
                    expPTS = 3 + ppg
                elif ppg - oppPts <= -2:
                    expPTS = 1.5 + ppg
                elif ppg - oppPts <= 0:
                    expPTS = 0 + ppg
                elif ppg - oppPts <= 2:
                    expPTS = -1.5 + ppg
                elif ppg - oppPts <= 4:
                    expPTS = -3 + ppg
                elif ppg - oppPts <= 6:
                    expPTS = -4.5 + ppg
                elif ppg - oppPts <= 8:
                    expPTS = -6 + ppg
                elif ppg - oppPts <= 10:
                    expPTS = -7.5 + ppg
                else:
                    expPTS = -9 + ppg

                #if loca == "Home":
                if ppgH - oppPts <= -4:
                    expPTSH += 3 + ppgH
                elif ppgH - oppPts <= -2:
                    expPTSH += 1.5 + ppgH
                elif ppgH - oppPts <= 0:
                    expPTSH += 0 + ppgH
                elif ppgH - oppPts <= 2:
                    expPTSH += -1.5 + ppgH
                elif ppgH - oppPts <= 4:
                    expPTSH += -3 + ppgH
                elif ppgH - oppPts <= 6:
                    expPTSH += -4.5 + ppgH
                elif ppgH - oppPts <= 8:
                    expPTSH += -6 + ppgH
                elif ppgH - oppPts <= 10:
                    expPTSH += -7.5 + ppgH
                else:
                    expPTSH += -9 + ppgH
                #else:
                if ppgA - oppPts <= -4:
                    expPTSA += 3 + ppgA
                elif ppgA - oppPts <= -2:
                    expPTSA += 1.5 + ppgA
                elif ppgA - oppPts <= 0:
                    expPTSA += 0 + ppgA
                elif ppgA - oppPts <= 2:
                    expPTSA += -1.5 + ppgA
                elif ppgA - oppPts <= 4:
                    expPTSA += -3 + ppgA
                elif ppgA - oppPts <= 6:
                    expPTSA += -4.5 + ppgA
                elif ppgA - oppPts <= 8:
                    expPTSA += -6 + ppgA
                elif ppgA - oppPts <= 10:
                    expPTSA += -7.5 + ppgA
                else:
                    expPTSA += -9 + ppgA
                avgPT3 = float(((opp3pm+pt3)/2)*3)
                avgPT2 = float(((opp2PTFGA*pt2a))*2)
                avgFT = float(oppFTA*fta)
                expPTS2 = (avgPT3 + avgPT2 + avgFT)

                projPTSH = (expPTS + expPTS2 + expPTSH)/3
                projPTSA = (expPTS + expPTS2 + expPTSA)/3
                print('')
                print('Home Projected Points is: ' + str(int(projPTSH)))
                print('Away Projected Points is: ' + str(int(projPTSA)))
                print('')

                break
            else:
                num_Play += 1
            
        except:
            break

#PlayerPropPTS()
First5()